from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass
from typing import Any
from xml.etree import ElementTree as ET

from scoring import ReportError


NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}
MAIN_NS = f"{{{NS['main']}}}"


ARCHIVE_10M_HEADERS = [
    "AssessmentYear",
    "ID",
    "id_10M",
    "PhotoHyperlink",
    "ReportCard",
    "Section",
    "Assessed",
    "Longitude",
    "Latitude",
    "Mangrove_Presence",
    "Naturalness",
    "Physical_Damage",
    "Modifier_Impact",
    "Notes",
]

ARCHIVE_50M_HEADERS = [
    "AssessmentYear",
    "SurveyDate",
    "Estuary",
    "Section",
    "ReportCard",
    "PointID",
    "Longitude",
    "Latitude",
    "PhotoHyperlink",
    "Assessed",
    "Mangrove_Presence_50m",
    "Density",
    "Maturity",
    "Condition_Score",
    "Connectivity",
    "Notes",
]

ASSESSMENT_BASE_HEADERS = [
    "AssessmentYear",
    "SurveyDate",
    "Estuary",
    "Section",
    "ReportCard",
    "PointID",
    "Longitude",
    "Latitude",
]


def _assessment_years(year: int) -> list[int]:
    return [year - 2, year - 1, year]


def _year_header(name: str, year: int | str) -> str:
    return f"{name}_{year}"


def _assessment_10m_headers(year: int) -> list[str]:
    years = _assessment_years(year)
    return (
        ASSESSMENT_BASE_HEADERS
        + [_year_header("PhotoHyperlink", item) for item in years]
        + [_year_header("Assessed", item) for item in years]
        + [_year_header("Mangrove_Presence", item) for item in years]
        + [_year_header("Naturalness", item) for item in years]
        + [_year_header("Physical_Damage", item) for item in years]
        + [_year_header("Modifier_Impact", item) for item in years]
        + [_year_header("Notes", year)]
    )


def _assessment_50m_headers(year: int) -> list[str]:
    years = _assessment_years(year)
    return (
        ASSESSMENT_BASE_HEADERS
        + [_year_header("PhotoHyperlink", item) for item in years]
        + [_year_header("Assessed", item) for item in years]
        + [_year_header("Mangrove_Presence_50m", item) for item in years]
        + [_year_header("Density", item) for item in years]
        + [_year_header("Maturity", item) for item in years]
        + [_year_header("Canopy_Cover", item) for item in years]
        + [_year_header("Connectivity", item) for item in years]
        + [_year_header("Notes", year)]
    )


@dataclass(frozen=True)
class LinkValue:
    display: str
    target: str


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ReportError(
            "The assessment workbook builder needs openpyxl. Run pip install -r requirements.txt and try again."
        ) from exc
    return Workbook, load_workbook, Font, PatternFill, get_column_letter


def _parse_assessment_year(value: Any) -> int:
    text = _clean(value)
    if not re.fullmatch(r"\d{4}", text):
        raise ReportError("Assessment year must be a four-digit year, for example 2025.")
    year = int(text)
    if year < 2000 or year > 2100:
        raise ReportError("Assessment year must be between 2000 and 2100.")
    return year


def _read_csv(file_bytes: bytes) -> list[dict[str, str]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [{key: _clean(value) for key, value in row.items()} for row in reader]


def _first_value(row: dict[str, str], *candidates: str) -> str:
    for candidate in candidates:
        if candidate in row and _clean(row.get(candidate)):
            return _clean(row.get(candidate))
    return ""


def _candidate_names(base: str, year: int) -> tuple[str, ...]:
    yy = str(year)[-2:]
    return (
        f"{base}{yy}",
        f"{base}_{yy}",
        f"{base}{year}",
        f"{base}_{year}",
        base,
    )


def _presence_50_candidates(year: int) -> tuple[str, ...]:
    yy = str(year)[-2:]
    return (
        f"Mangrove_Presence{yy}_50",
        f"Mangrove_Presence_{yy}_50",
        f"Mangrove_Presence{year}_50",
        f"Mangrove_Presence_{year}_50",
        "Mangrove_Presence_50m",
        "Mangrove_Presence_50",
        "Mangrove_Presence",
    )


def _hyperlink_candidates(year: int) -> tuple[str, ...]:
    yy = str(year)[-2:]
    return (
        f"Hyperlink{yy}",
        f"Hyperlink_{yy}",
        f"Hyperlink{year}",
        f"Hyperlink_{year}",
        "PhotoHyperlink",
        "Hyperlink",
        "Hyperlink2",
    )


def _normalize_headers(row: dict[str, str]) -> set[str]:
    return {_clean(value) for value in row.values() if _clean(value)}


def _looks_like_current_csv(rows: list[dict[str, str]]) -> bool:
    if not rows:
        return False
    headers = set(rows[0].keys())
    return bool(
        {"ID", "id_10m", "INT_50"}.issubset(headers)
        and ({"lon_10m_point", "lat_10m_point"}.issubset(headers) or {"lon_10m_p", "lat_10m_p"}.issubset(headers))
    )


def _raw_current_rows(file_bytes: bytes) -> list[dict[str, str]]:
    rows = _read_csv(file_bytes)
    if not _looks_like_current_csv(rows):
        raise ReportError(
            "Current-year CSV must include ID, id_10m, lon_10m_point, lat_10m_point, INT_50, Hyperlink, Section, and Assessed columns."
        )

    output = []
    for row in rows:
        point_id = _first_value(row, "id_10m", "id_10M", "PointID")
        if not point_id:
            continue
        output.append(
            {
                "ID": _first_value(row, "ID"),
                "PointID": point_id,
                "Longitude": _first_value(row, "lon_10m_point", "lon_10m_p", "Longitude"),
                "Latitude": _first_value(row, "lat_10m_point", "lat_10m_p", "Latitude"),
                "INT_50": _first_value(row, "INT_50"),
                "PhotoHyperlink": _first_value(row, "Hyperlink", "PhotoHyperlink"),
                "Section": _first_value(row, "Section"),
                "Assessed": _first_value(row, "Assessed"),
            }
        )
    if not output:
        raise ReportError("Current-year CSV did not contain any rows with an id_10m/PointID value.")
    return output


def _shared_strings(zf: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    return ["".join(text.text or "" for text in item.findall(".//main:t", NS)) for item in root.findall("main:si", NS)]


def _cell_value(cell: ET.Element, shared: list[str]) -> str:
    formula = cell.findtext("main:f", default="", namespaces=NS)
    value = cell.findtext("main:v", default="", namespaces=NS)
    cell_type = cell.attrib.get("t")
    if formula and "HYPERLINK" in formula.upper():
        return f"={formula}"
    if cell_type == "s" and value:
        return shared[int(value)]
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//main:t", NS))
    return value


def _column_index(cell_ref: str) -> int:
    column_name = "".join(ch for ch in cell_ref if ch.isalpha())
    total = 0
    for char in column_name.upper():
        total = total * 26 + ord(char) - 64
    return total


def _sheet_xml_path_by_name(zf: zipfile.ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    result: dict[str, str] = {}
    for sheet in workbook.find("main:sheets", NS):
        rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        target = rel_map[rel_id]
        if target.startswith("/"):
            path = target.lstrip("/")
        elif target.startswith("xl/"):
            path = target
        else:
            path = "xl/" + target
        result[sheet.attrib["name"]] = path
    return result


def _iter_sheet_rows(zf: zipfile.ZipFile, sheet_path: str, shared: list[str]):
    with zf.open(sheet_path) as sheet_file:
        for _, row in ET.iterparse(sheet_file, events=("end",)):
            if row.tag != f"{MAIN_NS}row":
                continue
            values: dict[int, str] = {}
            for cell in row.findall(f"{MAIN_NS}c"):
                column = _column_index(cell.attrib.get("r", ""))
                if column:
                    values[column] = _clean(_cell_value(cell, shared))
            if any(_clean(value) for value in values.values()):
                yield values
            row.clear()


def _workbook_records(zf: zipfile.ZipFile, sheet_path: str, shared: list[str]) -> list[dict[str, str]]:
    header_by_col: dict[int, str] | None = None
    records: list[dict[str, str]] = []
    for values in _iter_sheet_rows(zf, sheet_path, shared):
        if header_by_col is None:
            headers = {column: _clean(value) for column, value in values.items() if _clean(value)}
            normalized = set(headers.values())
            if (
                {"AssessmentYear", "id_10M"}.issubset(normalized)
                or {"AssessmentYear", "PointID"}.issubset(normalized)
                or {"id_10m", "lon_10m_point", "lat_10m_point"}.issubset(normalized)
            ):
                header_by_col = headers
            continue
        record = {header_by_col.get(column, f"Column{column}"): _clean(value) for column, value in values.items()}
        if any(record.values()):
            records.append(record)
    return records


def _year_from_sheet_name(sheet_name: str) -> int | None:
    match = re.search(r"(?<!\d)(20\d{2})(?!\d)", sheet_name)
    if not match:
        return None
    return int(match.group(1))


def _link_parts(value: str, for_50m: bool = False) -> LinkValue | str:
    text = _clean(value)
    if not text:
        return ""

    match = re.match(r'=HYPERLINK\("([^"]+)"\s*,\s*"([^"]*)"\)', text, flags=re.IGNORECASE)
    if match:
        target = match.group(1)
        display = match.group(2) or target.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    else:
        target = text
        display = text.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]

    if for_50m and "Frames_for_Analysis_50" not in target:
        target = target.replace("Frames_for_Analysis/", "Frames_for_Analysis_50/")
        target = target.replace("Frames_for_Analysis\\", "Frames_for_Analysis_50\\")

    return LinkValue(display=display, target=target)


def _plain_link_text(value: LinkValue | str) -> str:
    if isinstance(value, LinkValue):
        return value.target
    return _clean(value)


def _point_id_from_link(value: str) -> str:
    text = _clean(value)
    match = re.search(r"(?:^|[\\/])(\d+)_", text)
    if match:
        return match.group(1)
    match = re.match(r"(\d+)_", text)
    return match.group(1) if match else ""


def _record_key(record: dict[str, Any]) -> tuple[int, str] | None:
    year = _clean(record.get("AssessmentYear"))
    point_id = _clean(record.get("PointID") or record.get("id_10M") or record.get("id_10m"))
    if not year or not point_id:
        return None
    try:
        return int(float(year)), point_id
    except ValueError:
        return None


def _archive_10m_from_archive_row(row: dict[str, str]) -> dict[str, Any] | None:
    record = {header: _first_value(row, header) for header in ARCHIVE_10M_HEADERS}
    record["PointID"] = _first_value(row, "id_10M", "id_10m", "PointID")
    if not _record_key(record):
        return None
    return record


def _archive_50m_from_archive_row(row: dict[str, str]) -> dict[str, Any] | None:
    record = {header: _first_value(row, header) for header in ARCHIVE_50M_HEADERS}
    record["PointID"] = _first_value(row, "PointID", "id_10m", "id_10M")
    if not record["PointID"]:
        record["PointID"] = _point_id_from_link(record.get("PhotoHyperlink", ""))
    if not _record_key(record):
        return None
    return record


def _archive_10m_from_data_row(row: dict[str, str], year: int) -> dict[str, Any] | None:
    point_id = _first_value(row, "id_10m", "id_10M", "PointID")
    if not point_id:
        return None
    record = {
        "AssessmentYear": str(year),
        "ID": _first_value(row, "ID"),
        "id_10M": point_id,
        "PointID": point_id,
        "PhotoHyperlink": _first_value(row, *_hyperlink_candidates(year)),
        "ReportCard": _first_value(row, "ReportCard"),
        "Section": _first_value(row, "Section"),
        "Assessed": _first_value(row, *_candidate_names("Assessed", year)),
        "Longitude": _first_value(row, "lon_10m_point", "Longitude"),
        "Latitude": _first_value(row, "lat_10m_point", "Latitude"),
        "Mangrove_Presence": _first_value(row, *_candidate_names("Mangrove_Presence", year)),
        "Naturalness": _first_value(row, *_candidate_names("Naturalness", year)),
        "Physical_Damage": _first_value(row, *_candidate_names("Physical_Damage", year)),
        "Modifier_Impact": _first_value(row, "Modifier_Impact"),
        "Notes": _first_value(row, "Notes"),
    }
    return record


def _archive_50m_from_data_row(row: dict[str, str], year: int) -> dict[str, Any] | None:
    point_id = _first_value(row, "PointID", "id_10m", "id_10M")
    if not point_id:
        return None
    record = {
        "AssessmentYear": str(year),
        "SurveyDate": _first_value(row, "SurveyDate"),
        "Estuary": _first_value(row, "Estuary"),
        "Section": _first_value(row, "Section"),
        "ReportCard": _first_value(row, "ReportCard"),
        "PointID": point_id,
        "Longitude": _first_value(row, "lon_10m_point", "Longitude"),
        "Latitude": _first_value(row, "lat_10m_point", "Latitude"),
        "PhotoHyperlink": _first_value(row, *_hyperlink_candidates(year)),
        "Assessed": _first_value(row, *_candidate_names("Assessed", year)),
        "Mangrove_Presence_50m": _first_value(row, *_presence_50_candidates(year)),
        "Density": _first_value(row, *_candidate_names("Density", year)),
        "Maturity": _first_value(row, *_candidate_names("Maturity", year)),
        "Condition_Score": _first_value(row, *_candidate_names("Condition_Score", year)),
        "Connectivity": _first_value(row, *_candidate_names("Connectivity", year)),
        "Notes": _first_value(row, "Notes"),
    }
    return record


def _dedupe_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_key: dict[tuple[int, str], dict[str, Any]] = {}
    for record in records:
        key = _record_key(record)
        if key is None:
            continue
        by_key[key] = record
    return [by_key[key] for key in sorted(by_key, key=lambda item: (item[0], int(item[1]) if item[1].isdigit() else item[1]))]


def _load_archive_records(file_bytes: bytes, selected_year: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not file_bytes:
        return [], []
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            shared = _shared_strings(zf)
            sheet_map = _sheet_xml_path_by_name(zf)
            lower_names = {name.lower(): name for name in sheet_map}

            records_10m: list[dict[str, Any]] = []
            records_50m: list[dict[str, Any]] = []

            archive_10m_name = lower_names.get("archive_10m")
            archive_50m_name = lower_names.get("archive_50m")
            if archive_10m_name:
                for row in _workbook_records(zf, sheet_map[archive_10m_name], shared):
                    record = _archive_10m_from_archive_row(row)
                    if record is not None and str(record.get("AssessmentYear")) != str(selected_year):
                        records_10m.append(record)
            if archive_50m_name:
                for row in _workbook_records(zf, sheet_map[archive_50m_name], shared):
                    record = _archive_50m_from_archive_row(row)
                    if record is not None and str(record.get("AssessmentYear")) != str(selected_year):
                        records_50m.append(record)

            if archive_10m_name or archive_50m_name:
                return _dedupe_records(records_10m), _dedupe_records(records_50m)

            for sheet_name, sheet_path in sheet_map.items():
                sheet_year = _year_from_sheet_name(sheet_name)
                if sheet_year is None or sheet_year == selected_year:
                    continue
                rows = _workbook_records(zf, sheet_path, shared)
                if not rows:
                    continue
                is_50m_sheet = bool(re.search(r"(^|_)50($|_)", sheet_name, flags=re.IGNORECASE))
                for row in rows:
                    if is_50m_sheet:
                        record_50m = _archive_50m_from_data_row(row, sheet_year)
                        if record_50m is not None:
                            records_50m.append(record_50m)
                    else:
                        record_10m = _archive_10m_from_data_row(row, sheet_year)
                        if record_10m is not None:
                            records_10m.append(record_10m)
    except zipfile.BadZipFile as exc:
        raise ReportError("Archive workbook must be an .xlsx or .xlsm file.") from exc

    return _dedupe_records(records_10m), _dedupe_records(records_50m)


def _archive_lookup(records: list[dict[str, Any]]) -> dict[tuple[int, str], dict[str, Any]]:
    lookup = {}
    for record in records:
        key = _record_key(record)
        if key is not None:
            lookup[key] = record
    return lookup


def _should_prefill(current_assessed: str, previous: dict[str, Any] | None) -> bool:
    return current_assessed.upper() != "X" and bool(previous)


def _build_assessment_10m_rows(
    current_rows: list[dict[str, str]],
    archive_10m: list[dict[str, Any]],
    year: int,
    estuary: str,
) -> list[list[Any]]:
    prev = _archive_lookup(archive_10m)
    prev2_year = year - 2
    prev1_year = year - 1
    rows = []
    for current in current_rows:
        point_id = current["PointID"]
        prev1 = prev.get((prev1_year, point_id))
        prev2 = prev.get((prev2_year, point_id))
        current_assessed = current.get("Assessed", "")
        fill = _should_prefill(current_assessed, prev1)
        section = current.get("Section") or _clean((prev1 or {}).get("Section")) or _clean((prev1 or {}).get("ReportCard"))
        report_card = _clean((prev1 or {}).get("ReportCard")) or section
        current_mangrove_presence = _clean(prev1.get("Mangrove_Presence")) if fill else ""
        current_naturalness = _clean(prev1.get("Naturalness")) if fill else ""
        current_physical_damage = _clean(prev1.get("Physical_Damage")) if fill else ""
        current_modifier_impact = _clean(prev1.get("Modifier_Impact")) if fill else ""
        rows.append(
            [
                year,
                "",
                estuary,
                section,
                report_card,
                point_id,
                current.get("Longitude", ""),
                current.get("Latitude", ""),
                _link_parts(_clean((prev2 or {}).get("PhotoHyperlink"))),
                _link_parts(_clean((prev1 or {}).get("PhotoHyperlink"))),
                _link_parts(current.get("PhotoHyperlink", "")),
                _clean((prev2 or {}).get("Assessed")),
                _clean((prev1 or {}).get("Assessed")),
                current_assessed,
                _clean((prev2 or {}).get("Mangrove_Presence")),
                _clean((prev1 or {}).get("Mangrove_Presence")),
                current_mangrove_presence,
                _clean((prev2 or {}).get("Naturalness")),
                _clean((prev1 or {}).get("Naturalness")),
                current_naturalness,
                _clean((prev2 or {}).get("Physical_Damage")),
                _clean((prev1 or {}).get("Physical_Damage")),
                current_physical_damage,
                _clean((prev2 or {}).get("Modifier_Impact")),
                _clean((prev1 or {}).get("Modifier_Impact")),
                current_modifier_impact,
                "",
            ]
        )
    return rows


def _build_assessment_50m_rows(
    current_rows: list[dict[str, str]],
    archive_50m: list[dict[str, Any]],
    year: int,
    estuary: str,
) -> list[list[Any]]:
    prev = _archive_lookup(archive_50m)
    prev2_year = year - 2
    prev1_year = year - 1
    rows = []
    for current in current_rows:
        if not current.get("INT_50"):
            continue
        point_id = current["PointID"]
        prev1 = prev.get((prev1_year, point_id))
        prev2 = prev.get((prev2_year, point_id))
        current_assessed = current.get("Assessed", "")
        fill = _should_prefill(current_assessed, prev1)
        section = current.get("Section") or _clean((prev1 or {}).get("Section")) or _clean((prev1 or {}).get("ReportCard"))
        report_card = _clean((prev1 or {}).get("ReportCard")) or section
        current_presence = _clean(prev1.get("Mangrove_Presence_50m")) if fill else ""
        current_density = _clean(prev1.get("Density")) if fill else ""
        current_maturity = _clean(prev1.get("Maturity")) if fill else ""
        current_canopy_cover = _clean(prev1.get("Condition_Score")) if fill else ""
        current_connectivity = _clean(prev1.get("Connectivity")) if fill else ""
        rows.append(
            [
                year,
                "",
                estuary,
                section,
                report_card,
                point_id,
                current.get("Longitude", ""),
                current.get("Latitude", ""),
                _link_parts(_clean((prev2 or {}).get("PhotoHyperlink")), for_50m=True),
                _link_parts(_clean((prev1 or {}).get("PhotoHyperlink")), for_50m=True),
                _link_parts(current.get("PhotoHyperlink", ""), for_50m=True),
                _clean((prev2 or {}).get("Assessed")),
                _clean((prev1 or {}).get("Assessed")),
                current_assessed,
                _clean((prev2 or {}).get("Mangrove_Presence_50m")),
                _clean((prev1 or {}).get("Mangrove_Presence_50m")),
                current_presence,
                _clean((prev2 or {}).get("Density")),
                _clean((prev1 or {}).get("Density")),
                current_density,
                _clean((prev2 or {}).get("Maturity")),
                _clean((prev1 or {}).get("Maturity")),
                current_maturity,
                _clean((prev2 or {}).get("Condition_Score")),
                _clean((prev1 or {}).get("Condition_Score")),
                current_canopy_cover,
                _clean((prev2 or {}).get("Connectivity")),
                _clean((prev1 or {}).get("Connectivity")),
                current_connectivity,
                "",
            ]
        )
    return rows


def _archive_values(record: dict[str, Any], headers: list[str]) -> list[Any]:
    values = []
    for header in headers:
        if header == "PhotoHyperlink":
            values.append(_link_parts(_clean(record.get(header))))
        elif header == "id_10M":
            values.append(_clean(record.get("id_10M") or record.get("PointID")))
        else:
            values.append(_clean(record.get(header)))
    return values


def _append_row(ws: Any, values: list[Any]) -> None:
    row_number = ws.max_row + 1
    for column_number, value in enumerate(values, start=1):
        cell = ws.cell(row=row_number, column=column_number)
        if isinstance(value, LinkValue):
            cell.value = value.display
            if value.target:
                cell.hyperlink = value.target
                cell.style = "Hyperlink"
        else:
            cell.value = value


def _populate_sheet(ws: Any, headers: list[str], rows: list[list[Any]], *, header_fill: Any, header_font: Any) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in rows:
        _append_row(ws, row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells[:200]:
            max_length = max(max_length, len(_clean(cell.value)))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 34)


def _highlight_current_year_columns(ws: Any, year: int, fill: Any, header_font: Any) -> None:
    suffix = f"_{year}"
    for header_cell in ws[1]:
        if not _clean(header_cell.value).endswith(suffix):
            continue
        header_cell.fill = fill
        header_cell.font = header_font
        for row_number in range(2, ws.max_row + 1):
            ws.cell(row=row_number, column=header_cell.column).fill = fill


def build_assessment_workbook(
    current_csv_bytes: bytes,
    *,
    assessment_year: Any,
    archive_workbook_bytes: bytes,
    estuary_name: str = "",
) -> bytes:
    year = _parse_assessment_year(assessment_year)
    estuary = _clean(estuary_name)
    if not archive_workbook_bytes:
        raise ReportError("Archive workbook is required so the previous two years can be loaded.")
    current_rows = _raw_current_rows(current_csv_bytes)
    archive_10m, archive_50m = _load_archive_records(archive_workbook_bytes, year)
    assessment_10m = _build_assessment_10m_rows(current_rows, archive_10m, year, estuary)
    assessment_50m = _build_assessment_50m_rows(current_rows, archive_50m, year, estuary)

    Workbook, _, Font, PatternFill, _ = _load_openpyxl()
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F6D49")
    current_fill = PatternFill("solid", fgColor="DDF2E5")
    current_header_font = Font(bold=True, color="183127")

    archive_10m_ws = wb.active
    archive_10m_ws.title = "Archive_10m"
    _populate_sheet(
        archive_10m_ws,
        ARCHIVE_10M_HEADERS,
        [_archive_values(record, ARCHIVE_10M_HEADERS) for record in archive_10m],
        header_fill=header_fill,
        header_font=header_font,
    )

    archive_50m_ws = wb.create_sheet("Archive_50m")
    _populate_sheet(
        archive_50m_ws,
        ARCHIVE_50M_HEADERS,
        [_archive_values(record, ARCHIVE_50M_HEADERS) for record in archive_50m],
        header_fill=header_fill,
        header_font=header_font,
    )

    assessment_10m_ws = wb.create_sheet("Assessment_10m")
    _populate_sheet(
        assessment_10m_ws,
        _assessment_10m_headers(year),
        assessment_10m,
        header_fill=header_fill,
        header_font=header_font,
    )

    assessment_50m_ws = wb.create_sheet("Assessment_50m")
    _populate_sheet(
        assessment_50m_ws,
        _assessment_50m_headers(year),
        assessment_50m,
        header_fill=header_fill,
        header_font=header_font,
    )

    for ws in (assessment_10m_ws, assessment_50m_ws):
        _highlight_current_year_columns(ws, year, current_fill, current_header_font)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _sheet_header_map(ws: Any) -> dict[str, int]:
    return {_clean(cell.value): cell.column for cell in ws[1] if _clean(cell.value)}


def _row_value(ws: Any, row_number: int, headers: dict[str, int], name: str) -> str:
    column = headers.get(name)
    if not column:
        return ""
    return _clean(ws.cell(row=row_number, column=column).value)


def _row_value_any(ws: Any, row_number: int, headers: dict[str, int], *names: str) -> str:
    for name in names:
        value = _row_value(ws, row_number, headers, name)
        if value:
            return value
    return ""


def _row_link(ws: Any, row_number: int, headers: dict[str, int], name: str, *, for_50m: bool = False) -> LinkValue | str:
    column = headers.get(name)
    if not column:
        return ""
    cell = ws.cell(row=row_number, column=column)
    target = _clean(getattr(cell.hyperlink, "target", "")) if cell.hyperlink else ""
    display = _clean(cell.value)
    if target:
        return _link_parts(f'=HYPERLINK("{target}","{display or target}")', for_50m=for_50m)
    return _link_parts(display, for_50m=for_50m)


def _row_link_any(
    ws: Any,
    row_number: int,
    headers: dict[str, int],
    *names: str,
    for_50m: bool = False,
) -> LinkValue | str:
    for name in names:
        column = headers.get(name)
        if not column:
            continue
        link = _row_link(ws, row_number, headers, name, for_50m=for_50m)
        if _plain_link_text(link):
            return link
    return ""


def _append_year(ws: Any, headers: dict[str, int]) -> str:
    year = _row_value(ws, 2, headers, "AssessmentYear")
    if year:
        return year
    years = []
    for header in headers:
        for match in re.finditer(r"20\d{2}", header):
            years.append(match.group(0))
    return max(years) if years else ""


def _remove_archive_year_rows(ws: Any, year: str) -> None:
    headers = _sheet_header_map(ws)
    year_col = headers.get("AssessmentYear")
    if not year_col:
        return
    for row_number in range(ws.max_row, 1, -1):
        if _clean(ws.cell(row=row_number, column=year_col).value) == year:
            ws.delete_rows(row_number, 1)


def _append_assessment_10m_to_archive(wb: Any) -> int:
    if "Assessment_10m" not in wb.sheetnames or "Archive_10m" not in wb.sheetnames:
        return 0
    assessment = wb["Assessment_10m"]
    archive = wb["Archive_10m"]
    headers = _sheet_header_map(assessment)
    year = _append_year(assessment, headers)
    if not year:
        return 0
    _remove_archive_year_rows(archive, year)
    count = 0
    for row_number in range(2, assessment.max_row + 1):
        point_id = _row_value(assessment, row_number, headers, "PointID")
        if not point_id:
            continue
        row = [
            year,
            "",
            point_id,
            _row_link_any(assessment, row_number, headers, _year_header("PhotoHyperlink", year), "PhotoHyperlink"),
            _row_value(assessment, row_number, headers, "ReportCard"),
            _row_value(assessment, row_number, headers, "Section"),
            _row_value_any(assessment, row_number, headers, _year_header("Assessed", year), "Current_Assessed"),
            _row_value(assessment, row_number, headers, "Longitude"),
            _row_value(assessment, row_number, headers, "Latitude"),
            _row_value_any(
                assessment,
                row_number,
                headers,
                _year_header("Mangrove_Presence", year),
                "Current_Mangrove_Presence",
            ),
            _row_value_any(assessment, row_number, headers, _year_header("Naturalness", year), "Current_Naturalness"),
            _row_value_any(
                assessment,
                row_number,
                headers,
                _year_header("Physical_Damage", year),
                "Current_Physical_Damage",
            ),
            _row_value_any(
                assessment,
                row_number,
                headers,
                _year_header("Modifier_Impact", year),
                "Current_Modifier_Impact",
            ),
            _row_value_any(assessment, row_number, headers, _year_header("Notes", year), "Current_Notes"),
        ]
        _append_row(archive, row)
        count += 1
    return count


def _append_assessment_50m_to_archive(wb: Any) -> int:
    if "Assessment_50m" not in wb.sheetnames or "Archive_50m" not in wb.sheetnames:
        return 0
    assessment = wb["Assessment_50m"]
    archive = wb["Archive_50m"]
    headers = _sheet_header_map(assessment)
    year = _append_year(assessment, headers)
    if not year:
        return 0
    _remove_archive_year_rows(archive, year)
    count = 0
    for row_number in range(2, assessment.max_row + 1):
        point_id = _row_value(assessment, row_number, headers, "PointID")
        if not point_id:
            continue
        row = [
            year,
            _row_value(assessment, row_number, headers, "SurveyDate"),
            _row_value(assessment, row_number, headers, "Estuary"),
            _row_value(assessment, row_number, headers, "Section"),
            _row_value(assessment, row_number, headers, "ReportCard"),
            point_id,
            _row_value(assessment, row_number, headers, "Longitude"),
            _row_value(assessment, row_number, headers, "Latitude"),
            _row_link_any(
                assessment,
                row_number,
                headers,
                _year_header("PhotoHyperlink", year),
                "PhotoHyperlink",
                for_50m=True,
            ),
            _row_value_any(assessment, row_number, headers, _year_header("Assessed", year), "Current_Assessed"),
            _row_value_any(
                assessment,
                row_number,
                headers,
                _year_header("Mangrove_Presence_50m", year),
                "Current_Mangrove_Presence_50m",
            ),
            _row_value_any(assessment, row_number, headers, _year_header("Density", year), "Current_Density"),
            _row_value_any(assessment, row_number, headers, _year_header("Maturity", year), "Current_Maturity"),
            _row_value_any(
                assessment,
                row_number,
                headers,
                _year_header("Canopy_Cover", year),
                _year_header("Condition_Score", year),
                "Current_Condition_Score",
            ),
            _row_value_any(assessment, row_number, headers, _year_header("Connectivity", year), "Current_Connectivity"),
            _row_value_any(assessment, row_number, headers, _year_header("Notes", year), "Current_Notes"),
        ]
        _append_row(archive, row)
        count += 1
    return count


def workbook_has_assessment_sheets(file_bytes: bytes) -> bool:
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            sheet_names = {name.lower() for name in _sheet_xml_path_by_name(zf)}
    except zipfile.BadZipFile:
        return False
    return {"assessment_10m", "assessment_50m"}.issubset(sheet_names)


def append_assessment_to_archive_workbook(file_bytes: bytes) -> tuple[bytes, dict[str, int]]:
    _, load_workbook, _, _, _ = _load_openpyxl()
    try:
        wb = load_workbook(io.BytesIO(file_bytes))
    except Exception as exc:
        raise ReportError("Could not update the archive sheets in the assessment workbook.") from exc

    count_10m = _append_assessment_10m_to_archive(wb)
    count_50m = _append_assessment_50m_to_archive(wb)
    for sheet_name in ("Assessment_10m", "Assessment_50m"):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), {"archive_rows_10m": count_10m, "archive_rows_50m": count_50m}
